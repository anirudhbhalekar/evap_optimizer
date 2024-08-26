import numpy as np
import openpyxl

class EvapOptimizer():
    def __init__(self) -> None:

        self.init = True
        self.data_array = None
        self.row_count = 0
        self.file_name = None

        self.amb_temp_ind = None
        self.rh_ind = None
        self.evap_rate_ind = None
        self.heat_pow_ind = None
        self.fan_pow_ind = None

        self.max_stor_RH = -1
        self.max_stor_temp = -1
        self.min_stor_RH = 100
        self.min_stor_temp = 100

        self.target_per_tube = None
        self.energy_consumption = 0

        self.is_optimized = False
        self.optimal_config = []

    def load_file(self, file_name):
        self.file_name = file_name
        workbook = openpyxl.load_workbook(file_name)
        try:
            sheet = workbook["modif_database"]
            print("Database Found".center(100))
        except Exception:
            raise ValueError("Database sheet does not exist")

        self.row_count = sheet.max_row
        self.data_array = []

        for ind, cell in enumerate(sheet[1]):
            # find the indices for: ambient temp, ambient RH, evap rate and powers
            if "ambient temperature" in str(cell.value).lower():
                self.amb_temp_ind = ind
            elif "relative humidity (%)" in str(cell.value).lower():
                self.rh_ind = ind
            elif "evaporation rate" in str(cell.value).lower():
                self.evap_rate_ind = ind
            elif "heating power" in str(cell.value).lower():
                self.heat_pow_ind = ind
            elif "fan power" in str(cell.value).lower():
                self.fan_pow_ind = ind
            else:
                pass

        for row_ind in range(2, self.row_count + 1, 1):
            tmp_arr = []
            row = sheet[row_ind]

            if float(row[self.amb_temp_ind].value) > self.max_stor_temp: self.max_stor_temp = float(row[self.amb_temp_ind].value)
            if float(row[self.rh_ind].value) > self.max_stor_RH: self.max_stor_RH = float(row[self.rh_ind].value)
            if float(row[self.amb_temp_ind].value) < self.min_stor_temp: self.min_stor_temp = float(row[self.amb_temp_ind].value)
            if float(row[self.rh_ind].value) < self.min_stor_RH: self.min_stor_RH = float(row[self.rh_ind].value)
            for cell in row:
                tmp_arr.append(cell.value)
            self.data_array.append(tmp_arr)

    def load_forecast(self):
        # Method to load forecast
        if self.file_name is None:
            assert ValueError("Load Workbook First")
        workbook = openpyxl.load_workbook(self.file_name)
        try:
            sheet = workbook["forecast"]
            print("Database Found".center(100))
        except Exception:
            raise ValueError("Forecast sheet does not exist")

        this_row_count = sheet.max_row
        self.forecast_array = []
        self.time_array = []
        for row_ind in range(2, this_row_count + 1, 1):
            row = sheet[row_ind]
            try:
                self.forecast_array.append([int(row[1].value), int(row[2].value)])
                self.time_array.append(str(row[0].value))
            except:
                break

        # Load the target
        self.target_per_tube = float(sheet[2][3].value)

    def set_target(self, target):
        if self.file_name is None:
            assert ValueError("Load Workbook First")
        workbook = openpyxl.load_workbook(self.file_name)
        try:
            sheet = workbook["forecast"]
            print("Database Found".center(100))
        except Exception:
            raise ValueError("Forecast sheet does not exist")

        sheet.cell(2, 4).value = float(target)
        workbook.save(self.file_name)
        self.target_per_tube = target

    def initialise_arrays(self):
        # initialise power and evap rate array
        self.pow_arr = []
        self.evap_arr = []
        self.case_arr = []

        for [this_temp, this_rh] in self.forecast_array:
            closest_temp, closest_rh = 5 * round(int(this_temp)/5, 0), 5 * round(int(this_rh)/5, 0)

            if closest_temp > self.max_stor_temp: closest_temp = self.max_stor_temp
            if closest_rh > self.max_stor_RH: closest_rh = self.max_stor_RH
            if closest_temp < self.min_stor_temp: closest_temp = self.min_stor_temp
            if closest_rh < self.min_stor_RH: closest_rh = self.min_stor_RH

            tmp_pow_arr = []
            tmp_evap_arr = []
            tmp_case_arr = []

            for i, array in enumerate(self.data_array):
                if int(array[self.amb_temp_ind]) == closest_temp and int(array[self.rh_ind]) == closest_rh:
                    tmp_evap_arr.append(float(array[self.evap_rate_ind]))
                    tmp_pow_arr.append(float(array[self.heat_pow_ind]) + float(array[self.fan_pow_ind]))
                    tmp_case_arr.append(i)

            self.pow_arr.append(tmp_pow_arr)
            self.evap_arr.append(tmp_evap_arr)
            self.case_arr.append(tmp_case_arr)

    def optimizer(self):
        # We find the indices of the lowest energy possible be consumed
        # We look said indices up in the evap array to see if target is reached
        # If not - we find the lowest energy increment in each row - change the indice of said row to one that causes the smallest energy increment
        # Check again and repeat

        self.energy_consumption = 0
        self.is_optimized = False
        self.optimal_config = []


        indices = np.zeros(len(self.forecast_array))
        # Initialisation step
        for i, pow in enumerate(self.pow_arr):
            indices[i] = np.argmin(pow)

        self.prev_evap = -1
        self.current_evap = 0
        self.MAX_ITER = 10000
        count = 0

        # Iteration procedure
        while self.current_evap < self.target_per_tube and count < self.MAX_ITER:
            count += 1
            self.current_evap = 0
            for i, ind in enumerate(indices):
                self.current_evap += self.evap_arr[i][int(ind)] * 120 * 3
            
            for index, power_row in zip(indices, self.pow_arr):
                self.energy_consumption += 120 * 3 * power_row[int(index)] / 1000

            if self.current_evap > self.target_per_tube:
                self.is_optimized = True
                break 
            # If we are less than the target - we try to find the lowest energy increment possible
            min_energy_inc = np.inf
            new_indices = [-1,-1]

            for i, line in enumerate(self.pow_arr):
                curr_line_pow = line[int(indices[i])]
                for j, next_pow in enumerate(line):
                    diff = next_pow - curr_line_pow
                    if diff > 0 and diff < min_energy_inc:
                        min_energy_inc = diff
                        new_indices = [i, j]

            self.energy_consumption = 0
            indices[new_indices[0]] = new_indices[1]
            print("-"*100)
            print(f"Iteration {count}".center(100))
            print(f"Current Evap Rate : {self.current_evap}".center(100))
            print(f"Target Evap Rate : {self.target_per_tube}".center(100))
            print(f"Current Energy Consumption : {self.energy_consumption}".center(100))
            print("-"*100)

            if self.prev_evap == self.current_evap: 
                break 

            self.prev_evap = self.current_evap


        for i, index in enumerate(indices):
            optimal_case = self.case_arr[i][int(index)]
            optimal_case_config = self.data_array[int(optimal_case)]
            self.optimal_config.append(optimal_case_config)

    def store_results(self):
        # Load the results of the optimal configuration into excel workbook
        self.total_heating_power = 0
        self.total_fan_power = 0
        if self.file_name is None:
            assert ValueError("Load Workbook First")
        workbook = openpyxl.load_workbook(self.file_name)
        try:
            sheet = workbook["settings"]
            print("Database Found".center(100))
        except Exception:
            raise ValueError("Forecast sheet does not exist")

        for i, config in enumerate(self.optimal_config):
            row_ind = i + 2

            sheet.cell(row_ind, 4).value = config[0]
            sheet.cell(row_ind, 5).value = config[3]
            sheet.cell(row_ind, 6).value = config[4]
            sheet.cell(row_ind, 7).value = config[5]
            sheet.cell(row_ind, 8).value = config[9]
            sheet.cell(row_ind, 9).value = config[10]
            sheet.cell(row_ind, 10).value = config[11]
            sheet.cell(row_ind, 11).value = config[12]
            sheet.cell(row_ind, 12).value = config[13]
            sheet.cell(row_ind, 13).value = config[14]

            self.total_heating_power += float(config[13])
            self.total_fan_power += float(config[14])

        #Fan efficiency is 0.25 and there are 120 pans (3 hour intervals)
        self.total_fan_power *= 120 * 3 / 1000 / 0.25 
        self.total_heating_power *= 120 * 3 / 1000 
        
        workbook.save(self.file_name)

    def set_forecast(self, dt_arr, temp_arr, rh_arr):
        if self.file_name is None:
            assert ValueError("Load Workbook First")
        workbook = openpyxl.load_workbook(self.file_name)
        try:
            sheet = workbook["forecast"]
        except Exception:
            raise ValueError("Forecast sheet does not exist")
        
        for j in range(len(dt_arr)): 
            sheet.cell(j+2, 1).value = dt_arr[j]
            sheet.cell(j+2, 2).value = temp_arr[j]
            sheet.cell(j+2, 3).value = rh_arr[j]
        
        workbook.save(self.file_name)
        self.load_forecast()



        
